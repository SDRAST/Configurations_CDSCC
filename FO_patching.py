# -*- coding: utf-8 -*-
"""
Module to read F/O patching spreadsheet

This depends on the spreadsheet having the following format::
  Row  1 - Column name
  Row  2 - Selected patch; all blank except for the cell of the selected patch
  Row  3 - Band 18, Receiver 1, Pol E, IF L
  ...
  Row 42 - Band 26, Receiver 2, Pol H, IF U
"""
import os
import logging

import openpyxl
from openpyxl.reader.excel import InvalidFileException

import DatesTimes as DT

from support.excel import *

module_logger = logging.getLogger(__name__)

# repo_path = "/usr/local/lib/python2.7/DSN-Sci-packages/"
# modulepath = repo_path+"MonitorControl/Configurations/CDSCC/"
modulepath = os.path.dirname(os.path.abspath(__file__)) + "/"
paramfile = "FO_patching.xlsx"

label_map = {"E": 1, "H": 2, "L": 1, "U": 2}

class OpenpyxlCompatibility(object):

  def __init__(self):
    version_info = openpyxl.__version__
    major, minor, baby = list(map(int, version_info.split(".")))
    self.major = major
    self.minor = minor
    self.baby = baby
    self.version_info = [major, minor, baby]

  def cell(self, worksheet,row=0,column=0):
    """
    Enable compatibility with _newer_ releases of openpyxl.
    """
    if self.major > 1:
      row += 1
      column += 1
    return worksheet.cell(row=row, column=column)

  def get_sheet_names(self, workbook):
    if self.major > 1:
      return workbook.sheetnames
    else:
      return workbook.get_sheet_names()

  def get_sheet_by_name(self, workbook, name):
    if self.major > 1:
      return workbook[name]
    else:
      return workbook.get_sheet_by_name(name)

compat = OpenpyxlCompatibility()

class DistributionAssembly(object):
  """
  Public Attributes::
    logger      - logging.Logger instance
    paramfile   - name of CDSCC K-band patch panel spreadsheet
    parampath   - path to CDSCC configuration directory
    patching    - current IF patching dict
    patchname   - column title of current patching
    sheet_names - sorted names of sheets in patch panel spreadsheet (dates)
    workbook    - openpyxl.workbook.Workbook instance
  """
  def __init__(self, parampath=modulepath, paramfile=paramfile):
    """
    Create an instance of DistributionAssembly()
    """
    self.parampath = parampath
    self.paramfile = paramfile
    self.logger = logging.getLogger(module_logger.name+".DistributionAssembly")
    self.logger.debug("__init__: initializing")
    self._open_patchpanel_spreadsheet()
    self.patching = self.get_patching()

  def _open_patchpanel_spreadsheet(self):
    """
    Get the firmware summary worksheet
    """
    self.logger.debug("_open_patchpanel_spreadsheet: for %s",
      self.parampath+self.paramfile)
    try:
      param_file_path = os.path.join(self.parampath,self.paramfile)
      self.workbook = openpyxl.load_workbook(param_file_path)
    except IOError as details:
      self.logger.error(
      "_open_patchpanel_spreadsheet: loading spreadsheet failed with IO error.",
                        exc_info=True)
      raise IOError
    except InvalidFileException:
      self.logger.error(
      "_open_patchpanel_spreadsheet: .reader.excel doesn't like this file.",
                        exc_info=True)
      raise InvalidFileException
    except AttributeError:
      self.logger.error(
                        "_open_patchpanel_spreadsheet: attribute error.",
                        exc_info=True)
      raise AttributeError
    # get the current worksheet
    self.sheet_names = compat.get_sheet_names(self.workbook) #.get_sheet_names()
    self.sheet_names.sort()
    self.logger.debug("_open_patchpanel_spreadsheet: sheet names: %s",
                      str(self.sheet_names))
    # last worksheet by date YEAR/DOY
    self.worksheet = compat.get_sheet_by_name(self.workbook, self.sheet_names[-1])
    column_numbers = support.excel.get_column_names(self.worksheet)
    self.logger.debug("_open_patchpanel_spreadsheet: columns found in %s:",
                      self.sheet_names[-1])
    # find the current patching
    for number in list(column_numbers.keys()):
      if column_numbers[number]:
        self.logger.debug("_open_patchpanel_spreadsheet: %s: %s",
                          number ,column_numbers[number])
    self.current_patch()
    self.logger.debug("_open_patchpanel_spreadsheet: current patch is %s",
                      self.patchname)
    self.column = support.excel.get_column_id(self.worksheet, self.patchname)
    self.logger.debug("_open_patchpanel_spreadsheet: active column is %s", 
                      self.column)
    

  def current_patch(self):
    """
    Find the patching currently in effect.

    If no patching is currently known, it steps through columns E through I of
    row 2 (index 1) until it finds a non-empty cell.
    """
    self.patchname = None
    for column_idx in range(1,11):
      self.logger.debug("current_patch: checking column index %d", column_idx)
      cell_value = compat.cell(self.worksheet, row=1, 
                                     column=column_idx).value
      self.logger.debug("current_patch: cell value = %s", cell_value)
      if cell_value:
        current = compat.cell(self.worksheet, row=1, 
                                            column=column_idx).value
        #self.logger.debug("current_patch: found {}".format(current.encode("utf-8")))
        self.logger.debug("current_patch: found {}".format(current))
        if self.patchname:
          self.logger.error("current_patch: ambiguity: {} or {}".format(
              self.patchname, compat.cell(self.worksheet, row=1, 
                                                      column=column_idx).value))
          raise RuntimeException("patch ambiguity")
        else:
          # self.patchname = self.worksheet.cell(row=0,column=column).value
          self.patchname = compat.cell(self.worksheet,row=0,column=column_idx).value
          break
      else:
        pass
    return self.patchname

  def get(self, column_name, row):
    """
    Returns value for column name in the row, including merged cells
    """
    column = support.excel.get_column_id(self.worksheet, column_name)-OPENPYXL_INDEX
    column_data = support.excel.get_column(self.worksheet, column_name)
    while row > 0:
      if compat.cell(self.worksheet, row=row, column=column).value:
        return compat.cell(self.worksheet, row=row, column=column).value
      else:
        row -= 1
    return None

  def get_sheet_by_date(self, obsdate=None):
    """
    """
    if obsdate:
      if '/' in obsdate:
        parts = obsdate.split("/")
      elif '-' in obsdate:
        parts = obsdate.split("-")
      else:
        raise RuntimeWarning("get_sheet_by_date: date %s not recognized")
        return None
      year = parts[0]
      if len(parts) == 2:
        doy = parts[1]
      elif len(parts) == 3:
        doy = '%03d' % DT.day_of_year(int(parts[0]),
                                      int(parts[1]), int(parts[2]))
      for name in self.sheet_names:
        patchyear, patchdoy = name.split('-')
        if year != patchyear:
          continue
        if int(doy) > int(patchdoy):
          # try the next one
          sheetfound = name
        elif int(doy) < int(patchdoy):
          # passed the desired sheet
          self.worksheet = compat.get_sheet_by_name(self.workbook, name)
          return name
      # end name loop; none found
    # none found or no date given
    self.worksheet = compat.get_sheet_by_name(self.workbook, self.sheet_names[-1])
    return self.sheet_names[-1]

  def get_patching(self, obsdate=None):
    """
    Returns patching on the current or given date

    @param obsdate - "YYYY/DDD" or "YYYY/MM/DD"
    """
    IF_channel = {}
    for IF in range(1,17):
      rx_chan = {}
      self.logger.debug("get_patching: checking IF %d", IF)
      row = get_row_number(self.worksheet, self.column, IF)
      self.logger.debug("get_patching: IF %s is in row %s", IF, row)
      for item in ["Band", "Receiver", "Pol", "IF"]:
        value = self.get(item, row)
        if value == None:
          self.logger.error("get_patching: no %s for row %s", item, row)
        else:
          rx_chan[item] = value
      IF_channel[IF] = rx_chan
    return IF_channel

  def report_patching(self):
    IF_report = {}
    for IF in list(self.patching.keys()):
      rx_chan = self.patching[IF]
      IF_report[IF] = "R"+str(rx_chan["Receiver"]) \
                     +"-"+str(rx_chan["Band"]) \
                     +"-"+str(rx_chan["Pol"]) \
                     +"-IF"+str(label_map[str(rx_chan["IF"])])
    return IF_report

  def get_signals(self, device):
    """
    Returns the signals into the specified device.

    Currently known devices::
      'Power Meter' - (four) Hewlett Packard power meters
      'Radiometer'  - eight-head Date! power meter assembly
      'ROACH1'      - SAO spectrometer
      'ROACH2'      - GCPS spectrometer
    """
    try:
      inputs = support.excel.get_column(self.worksheet, device)[1:]
    except TypeError:
      self.logger.error("get_signals: device %s is not known", device)
      raise RuntimeError("device %s is not known; check capitalization." % device)
    inputs = support.excel.get_column(self.worksheet, device)[1:]
    self.logger.debug("get_signals: Column '%s' values: %s", device, inputs)
    sig_props = {}
    for index in range(len(inputs)):
      # row 2 is labelled 3 in the spreadsheet (for openpyxl 1.x)
      row = index+2
      channel_ID = inputs[index]
      self.logger.debug("get_signals: row %d input is %s", row, channel_ID)
      if inputs[index]:
        sig_props[channel_ID] = {}
        for item in ["Band", "Receiver", "Pol", "IF"]:
          value= self.get(item, row)
          if value == None:
            self.logger.error("get_signals: no %s for row %d", item, row)
          else:
            sig_props[channel_ID][item] = value
    return sig_props

  def get_inputs(self, device):
    """
    """
    try:
      inputs = support.excel.get_column(self.worksheet, device)[1:]
    except TypeError:
      self.logger.error("get_inputs: device %s is not known", device)
      raise RuntimeError("device %s is not known; check capitalization." % device)
    self.logger.debug("get_inputs: Column '%s' values: %s", device, inputs)
    channels = {}
    for index in range(len(inputs)):
      # row 2 is labelled 3 in the spreadsheet (for openpyxl 1.x)
      row = index+2
      channel_ID = inputs[index]
      self.logger.debug("get_inputs: row %d input is %s", row, channel_ID)
      if inputs[index]:
        channels[channel_ID] = {}
        channels[channel_ID]['RF'] = 'R'+str(self.get('Receiver',row)) \
                                    +'-'+str(self.get('Band',row)) \
                                    +str(self.get('Pol',row))
        channels[channel_ID]['IF'] = self.get('IF',row)
    return channels
